import re
from pathlib import Path

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("pandas is required. Install with: pip install pandas") from exc

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from exc


ROOT = Path(__file__).parent
CSV_PATH = ROOT / "results" / "batch_resultados_final.csv"
TEMPLATE_PATH = ROOT / "DatosPruebas2026_1.xlsx"
OUT_PATH = ROOT / "results" / "batch_resultados_final_organizado.xlsx"
TARGET_K = [2, 3, 4, 5]
ESTRATEGIA_MAP = {"KQNodes": "QNodes", "KGeoMIP": "Geometric"}
ERROR_PATTERN = re.compile(r"(memoryerror|recursionerror|error)", re.I)


def normalize_partition(value):
    if pd.isna(value):
        return "N/A"
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text if text else "N/A"


def canonical_column_names(df):
    lower = {col.lower(): col for col in df.columns}
    mapping = {}
    for key in ["sample", "k", "estrategia", "delta", "tiempo", "particion", "alcance", "mecanismo", "error"]:
        if key in lower:
            mapping[lower[key]] = key
    required = ["sample", "k", "estrategia", "delta", "tiempo", "alcance", "mecanismo"]
    missing = [key for key in required if key not in mapping.values()]
    if missing:
        raise SystemExit(f"Faltan columnas requeridas en el CSV: {missing}")
    return df.rename(columns=mapping)


def normalize_dataframe():
    if not CSV_PATH.exists():
        raise FileNotFoundError(f"CSV no encontrado: {CSV_PATH}")

    df = pd.read_csv(CSV_PATH, dtype=str, encoding="utf-8")
    df = canonical_column_names(df)
    df["sample"] = df["sample"].astype(str).str.replace(r"\.csv$", "", regex=True).str.strip()
    df["estrategia"] = (
        df["estrategia"].astype(str).str.strip().map(ESTRATEGIA_MAP).fillna(df["estrategia"].astype(str).str.strip())
    )
    df["k"] = pd.to_numeric(df["k"], errors="coerce").astype("Int64")
    df["alcance"] = df["alcance"].astype(str).fillna("").str.strip()
    df["mecanismo"] = df["mecanismo"].astype(str).fillna("").str.strip()
    df["particion"] = df.get("particion", pd.Series(["N/A"] * len(df))).apply(normalize_partition)
    df["delta"] = df["delta"].astype(str).str.strip()
    df["tiempo"] = df["tiempo"].astype(str).str.strip()
    if "error" in df.columns:
        df["error"] = df["error"].astype(str).str.strip()
    return df


def row_has_error(row):
    for col in ["sample", "estrategia", "delta", "tiempo", "particion", "alcance", "mecanismo"]:
        value = row.get(col, "")
        if isinstance(value, str) and ERROR_PATTERN.search(value):
            return True
    if "error" in row and isinstance(row["error"], str) and row["error"].strip():
        return True
    return False


def sanitize_error_value(value):
    if pd.isna(value):
        return "N/A"
    text = str(value).strip()
    if ERROR_PATTERN.search(text):
        return "N/A"
    return text


def clean_sample_rows(df):
    if df.empty:
        return df

    if "error" in df.columns:
        df = df.loc[~df["error"].astype(str).str.strip().astype(bool)].copy()

    if df.empty:
        return df

    for col in ["delta", "tiempo", "particion", "alcance", "mecanismo"]:
        if col in df.columns:
            df[col] = df[col].apply(sanitize_error_value)

    return df


def pivot_sample(sample_df):
    sample_df = clean_sample_rows(sample_df)
    if sample_df.empty:
        return None

    sample_df = sample_df[sample_df["estrategia"].isin(["QNodes", "Geometric"])]
    sample_df = sample_df[sample_df["k"].notna() & sample_df["k"].isin(TARGET_K)]
    if sample_df.empty:
        return None

    pivot = sample_df.pivot_table(
        index=["sample", "alcance", "mecanismo"],
        columns=["k", "estrategia"],
        values=["particion", "delta", "tiempo"],
        aggfunc="first",
    )

    ordered_columns = []
    for k in TARGET_K:
        for estrategia in ["QNodes", "Geometric"]:
            ordered_columns.extend(
                [("particion", k, estrategia), ("delta", k, estrategia), ("tiempo", k, estrategia)]
            )

    pivot = pivot.reindex(columns=pd.MultiIndex.from_tuples(ordered_columns))
    pivot = pivot.fillna("N/A")

    flat_columns = [f"k{k}_{estrategia}_{measure}" for measure, k, estrategia in pivot.columns]
    pivot.columns = flat_columns
    result = pivot.reset_index()
    result.insert(0, "#Prueba", range(1, len(result) + 1))
    return result


def build_result_frames():
    df = normalize_dataframe()
    samples = sorted(df["sample"].dropna().unique())
    results = {}

    for sample in samples:
        sample_df = df[df["sample"] == sample].copy()
        try:
            result = pivot_sample(sample_df)
            if result is None:
                result = pd.DataFrame(columns=["#Prueba", "sample", "alcance", "mecanismo"] + [f"k{k}_{estr}_{measure}" for k in TARGET_K for estr in ["QNodes", "Geometric"] for measure in ["particion", "delta", "tiempo"]])
                result.insert(1, "sample", sample)
            results[sample] = result
        except (MemoryError, RecursionError) as exc:
            print(f"Advertencia: error al procesar la red {sample}: {exc}")
            result = pd.DataFrame(columns=["#Prueba", "sample", "alcance", "mecanismo"] + [f"k{k}_{estr}_{measure}" for k in TARGET_K for estr in ["QNodes", "Geometric"] for measure in ["particion", "delta", "tiempo"]])
            result.insert(1, "sample", sample)
            results[sample] = result

    if not results:
        raise SystemExit("No se encontraron redes válidas para procesar.")
    return results


def find_header_row(ws, header_names):
    for row in range(1, 11):
        values = [str(ws.cell(row=row, column=col).value).strip().lower() if ws.cell(row=row, column=col).value else "" for col in range(1, ws.max_column + 1)]
        if all(name.lower() in values for name in header_names[:3]):
            return row
    return None


def write_sheet(ws, df):
    header = list(df.columns)
    header_row = find_header_row(ws, ["#Prueba", "Alcance", "Mecanismo"])
    if header_row is None:
        header_row = 1
        for col_idx, col_name in enumerate(header, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    start_row = header_row + 1

    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row.tolist(), start=1):
            ws.cell(row=start_row + r_idx, column=c_idx, value=value)

    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    if len(header) >= 4:
        ws.freeze_panes = ws.cell(row=start_row, column=4)


def write_workbook(result_frames):
    if TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
    else:
        wb = Workbook()
        if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active.cell(1, 1).value is None:
            wb.remove(wb.active)

    for sample, frame in result_frames.items():
        sheet_name = str(sample)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, frame.drop(columns=["sample"]))

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Archivo guardado en: {OUT_PATH}")


def main():
    print(f"Leyendo CSV: {CSV_PATH}")
    result_frames = build_result_frames()
    write_workbook(result_frames)


if __name__ == "__main__":
    main()
